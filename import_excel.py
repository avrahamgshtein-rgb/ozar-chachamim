#!/usr/bin/env python3
"""
Import sages from Excel file with Spotify queries
"""
import json
import openpyxl
import sys

# Hebrew support
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_FILE = "site-data/חכמי ישראל.xlsx"
OUTPUT_FILE = "data.json"

print("📖 Importing from Excel...")

try:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    print(f"✓ Opened: {EXCEL_FILE}")
    print(f"✓ Sheet name: {ws.title}")
    print(f"✓ Rows: {ws.max_row}")
    print(f"✓ Columns: {ws.max_column}")

    # Print first row to see headers
    print("\n📋 Column Headers:")
    headers = []
    for col_idx, cell in enumerate(ws[1], 1):
        value = cell.value
        headers.append(value)
        print(f"  Col {col_idx}: {value}")

    print("\n✅ Ready to import!")
    print(f"\nNow tell me:")
    print("1. Which column is the sage name? (e.g., 1)")
    print("2. Which column is the era/period? (e.g., 2)")
    print("3. Which column is the location? (e.g., 3)")
    print("4. Which column is Spotify query? (e.g., 4)")

except Exception as e:
    print(f"❌ Error: {e}")
    print(f"Make sure Excel file exists at: {EXCEL_FILE}")

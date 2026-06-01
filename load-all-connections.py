#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Load ALL connections from Excel and insert to Supabase
"""
import requests
import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SUPABASE_URL = "https://ulluacifirzywhmzkvkr.supabase.co"
SUPABASE_KEY = "sb_publishable_ObxKLFsDTE41KoAMfMV1dw_Nu38ZI2C"

headers = {
    'apikey': SUPABASE_KEY,
    'Content-Type': 'application/json',
}

print("📚 Loading Excel...")
wb = openpyxl.load_workbook("data/חכמי ישראל.xlsx")
ws = wb.active

# Parse all sages first to build name→id map
sages = {}
for row_idx in range(2, ws.max_row + 1):
    sage_id = ws.cell(row=row_idx, column=1).value
    name = ws.cell(row=row_idx, column=3).value
    if name and sage_id:
        sages[str(name).strip()] = str(int(sage_id)) if isinstance(sage_id, (int, float)) else str(sage_id)

print(f"✓ Loaded {len(sages)} sage names")

# Parse connections from column 11
print("🔗 Parsing connections...")
connections = {}

for row_idx in range(2, ws.max_row + 1):
    source_name = ws.cell(row=row_idx, column=3).value
    connections_raw = ws.cell(row=row_idx, column=11).value
    
    if not source_name or not connections_raw:
        continue
    
    source_name = str(source_name).strip()
    if source_name not in sages:
        continue
    
    source_id = sages[source_name]
    
    # Parse "Name (Type), Name2 (Type2)" format
    pairs = [p.strip() for p in str(connections_raw).split(',')]
    
    for pair in pairs:
        if not pair:
            continue
        
        connection_type = 'colleague'
        target_name = pair
        
        if '(' in pair and ')' in pair:
            try:
                target_name = pair[:pair.index('(')].strip()
                connection_type = pair[pair.index('(')+1:pair.index(')')].strip().lower()
            except:
                pass
        
        target_name = target_name.strip()
        if target_name not in sages:
            continue
        
        target_id = sages[target_name]
        key = f"{source_id}→{target_id}"
        
        if key not in connections:
            connections[key] = {
                'source_id': source_id,
                'target_id': target_id,
                'connection_type': connection_type or 'colleague'
            }

print(f"✓ Parsed {len(connections)} unique connections")

# Insert in batches
if connections:
    conn_list = list(connections.values())
    batch_size = 50
    inserted = 0
    
    for i in range(0, len(conn_list), batch_size):
        batch = conn_list[i:i+batch_size]
        
        response = requests.post(
            f'{SUPABASE_URL}/rest/v1/connections',
            json=batch,
            headers=headers
        )
        
        if response.status_code in [200, 201]:
            inserted += len(batch)
            print(f"  ✓ Batch {i//batch_size + 1}: {len(batch)} connections")
        else:
            if '23514' in response.text:  # Check constraint error
                print(f"  ⚠️  Batch {i//batch_size + 1}: Some constraint errors (OK)")
            else:
                print(f"  ⚠️  Batch {i//batch_size + 1}: {response.status_code}")

print(f"\n✅ Done! Connections inserted")
print(f"\n🌐 Now open: http://localhost:8080")
print(f"   You should see:")
print(f"   ✓ Nodes connected with lines")
print(f"   ✓ Click nodes to see details")

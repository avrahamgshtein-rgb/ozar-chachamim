#!/usr/bin/env python3
"""
Extract relationship edges from Excel "דמויות קשורות" column
and generate links for the D3 graph
"""
import json
import openpyxl
import sys
import io
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_FILE = "data/חכמי ישראל.xlsx"
DATA_FILE = "data.json"

print("=" * 100)
print("🔗 GENERATING RELATIONSHIP EDGES FROM EXCEL")
print("=" * 100)

try:
    # Load Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Build name→id mapping
    name_to_id = {}
    id_to_name = {}

    for row_idx in range(2, ws.max_row + 1):
        sage_id = ws.cell(row=row_idx, column=1).value
        name = ws.cell(row=row_idx, column=3).value

        if name and str(name).strip():
            name_str = str(name).strip()
            name_to_id[name_str] = str(sage_id)
            id_to_name[str(sage_id)] = name_str

    print(f"\n✓ Mapped {len(name_to_id)} sages from Excel")

    # Extract relationships
    links = []
    relationships_found = defaultdict(int)

    for row_idx in range(2, ws.max_row + 1):
        sage_id = ws.cell(row=row_idx, column=1).value
        name = ws.cell(row=row_idx, column=3).value
        related_raw = ws.cell(row=row_idx, column=11).value  # Col 11: דמויות קשורות

        if not name or not related_raw:
            continue

        name_str = str(name).strip()
        sage_id_str = str(sage_id) if sage_id else ""
        related_text = str(related_raw).strip()

        # Parse related sages (format: "Name1 (type), Name2 (type), ...")
        if related_text and related_text != 'nan':
            # Split by comma
            pairs = [p.strip() for p in related_text.split(',')]

            for pair in pairs:
                if not pair:
                    continue

                # Try to extract name and type: "Name (Type)" or just "Name"
                type_val = 'colleague'  # default
                target_name = pair

                if '(' in pair and ')' in pair:
                    try:
                        target_name = pair[:pair.index('(')].strip()
                        type_val = pair[pair.index('(')+1:pair.index(')')].strip().lower()
                    except:
                        target_name = pair

                # Map type to link category
                type_map = {
                    'student': 'student',
                    'תלמיד': 'student',
                    'teacher': 'influence',
                    'רב': 'influence',
                    'master': 'influence',
                    'influence': 'influence',
                    'influenced': 'influence',
                    'opponent': 'oppose',
                    'oppose': 'oppose',
                    'colleague': 'colleague',
                    'contemporary': 'colleague',
                    'predecessor': 'predecessor',
                    'precursor': 'predecessor',
                }
                link_type = type_map.get(type_val, 'colleague')

                # Find target ID
                if target_name in name_to_id:
                    target_id = name_to_id[target_name]

                    # Add link (avoid duplicates)
                    link = {
                        'source': sage_id_str,
                        'target': target_id,
                        'type': link_type
                    }

                    # Check if not duplicate
                    is_duplicate = False
                    for existing in links:
                        if (existing['source'] == link['source'] and
                            existing['target'] == link['target'] and
                            existing['type'] == link['type']):
                            is_duplicate = True
                            break

                    if not is_duplicate:
                        links.append(link)
                        relationships_found[link_type] += 1

    print(f"\n✓ Extracted {len(links)} relationship edges:")
    for rel_type, count in sorted(relationships_found.items(), key=lambda x: -x[1]):
        print(f"   • {rel_type}: {count}")

    # Load data.json and add links
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)

    data['links'] = links

    # Save updated data
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Updated {DATA_FILE} with {len(links)} edges")

    # Stats
    print(f"\n📊 Edge Statistics:")
    print(f"   • Nodes: {len(data['nodes'])}")
    print(f"   • Links: {len(data['links'])}")
    print(f"   • Graph density: {len(links) / (len(data['nodes']) * (len(data['nodes'])-1) / 2) * 100:.2f}%")

except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()

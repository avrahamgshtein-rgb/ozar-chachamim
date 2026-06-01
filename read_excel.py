import openpyxl
import json
import sys
import io

# Fix UTF-8 encoding for output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('site-data/חכמי ישראל.xlsx')
ws = wb.active

# Get headers
headers = []
for col in range(1, ws.max_column + 1):
    val = ws.cell(1, col).value
    if val:
        headers.append(val)

print("COLUMNS:")
for i, h in enumerate(headers, 1):
    print(f"{i}: {h}")

print(f"\nTotal data rows: {ws.max_row - 1}")

print("\nFirst sage data:")
row_data = {}
for col_idx, header in enumerate(headers, 1):
    val = ws.cell(2, col_idx).value
    row_data[header] = val
print(json.dumps(row_data, ensure_ascii=False, indent=2))

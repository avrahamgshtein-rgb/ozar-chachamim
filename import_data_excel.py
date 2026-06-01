#!/usr/bin/env python3
"""
Import sages from data/ folder Excel with Spotify queries
"""
import json
import openpyxl
import sys
import io
import os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_FILE = "data/חכמי ישראל.xlsx"
OUTPUT_FILE = "data.json"

print("📖 Importing from data/חכמי ישראל.xlsx...")

try:
    # Check if file exists
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ File not found: {EXCEL_FILE}")
        # Try alternate paths
        for path in ["data\\חכמי ישראל.xlsx", "./data/חכמי ישראל.xlsx"]:
            if os.path.exists(path):
                EXCEL_FILE = path
                print(f"✓ Found at: {path}")
                break

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    print(f"✓ Sheet: {ws.title}")
    print(f"✓ Max rows: {ws.max_row}")

    # Read headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    print("\n📋 Columns:")
    for idx, header in enumerate(headers[:12], 1):
        print(f"  {idx}: {header}")

    nodes = []
    skipped = 0

    # Import sages (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        try:
            sage_id = row[0]  # Col 1: ID
            chapter_type = row[1]  # Col 2: סוג פרק
            name = row[2]  # Col 3: שם הדמות
            period_years = row[3]  # Col 4: שנים
            location = row[4]  # Col 5: אזור
            era = row[5]  # Col 6: תקופה
            field = row[6]  # Col 7: תחום
            tags = row[7]  # Col 8: תגיות
            bio = row[8]  # Col 9: תקציר
            main_idea = row[9]  # Col 10: רעיון
            related = row[10]  # Col 11: דמויות קשורות
            spotify = row[11]  # Col 12: קישור ספוטיפיי

            # Filter: only add if has name
            if name and name.strip():
                # Create sage node
                node = {
                    "id": str(sage_id) if sage_id else str(len(nodes) + 1),
                    "label": str(name).strip(),
                    "group": (str(era).lower().replace(' ', '-').replace('/', '-') if era else 'unknown'),
                    "era": str(era) if era else 'Unknown',
                    "period": str(period_years) if period_years else '',
                    "location": str(location).strip() if location else '',
                    "field": str(field).strip() if field else 'Other',
                    "bio": str(bio).strip() if bio else f"Sage from {era}",
                    "spotifyQuery": str(spotify).strip() if spotify and str(spotify).strip() else name
                }
                nodes.append(node)
            else:
                skipped += 1

        except Exception as e:
            skipped += 1

    print(f"\n✓ Imported {len(nodes)} sages")
    print(f"⚠️  Skipped {skipped} rows")

    # Save
    data = {"nodes": nodes, "links": []}

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Saved to {OUTPUT_FILE}")
    print(f"\n📊 Sample sages:")
    for node in nodes[:5]:
        spotify_preview = node['spotifyQuery'][:40] + '...' if len(node['spotifyQuery']) > 40 else node['spotifyQuery']
        print(f"  • {node['label']} ({node['era']})")
        print(f"    Spotify: {spotify_preview}")

    print(f"\n🎵 Total with Spotify: {sum(1 for n in nodes if n.get('spotifyQuery'))}/{len(nodes)}")

except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()

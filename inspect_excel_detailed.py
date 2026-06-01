#!/usr/bin/env python3
"""
Detailed Excel inspection - read ALL columns and first 10 rows
"""
import openpyxl
import json
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_FILE = "data/חכמי ישראל.xlsx"

print("=" * 100)
print("📊 DETAILED EXCEL INSPECTION")
print("=" * 100)

try:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    print(f"\n📄 File: {EXCEL_FILE}")
    print(f"📋 Sheet: {ws.title}")
    print(f"📐 Dimensions: {ws.max_row} rows × {ws.max_column} columns")

    # Print ALL column headers with their numbers
    print("\n" + "=" * 100)
    print("📌 COLUMN HEADERS:")
    print("=" * 100)

    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        value = cell.value
        headers[col_idx] = value
        print(f"  Col {col_idx:2d}: {value}")

    # Print first 10 data rows in detail
    print("\n" + "=" * 100)
    print("📖 FIRST 10 SAMPLE ROWS (Row 2-11):")
    print("=" * 100)

    for row_idx in range(2, min(12, ws.max_row + 1)):
        print(f"\n🔹 Row {row_idx}:")
        row = ws[row_idx]
        for col_idx, cell in enumerate(row[:ws.max_column], 1):
            header = headers.get(col_idx, "Unknown")
            value = cell.value

            # Truncate long values for display
            if value:
                display_value = str(value)[:80]
                if len(str(value)) > 80:
                    display_value += "..."
            else:
                display_value = "[EMPTY]"

            print(f"    Col {col_idx:2d} ({header}): {display_value}")

    # Analyze data patterns
    print("\n" + "=" * 100)
    print("🔍 DATA PATTERNS:")
    print("=" * 100)

    # Check for Spotify URLs
    spotify_col = None
    for col_idx in range(1, ws.max_column + 1):
        header = headers.get(col_idx, "").lower()
        if "spotify" in header or "קישור" in headers.get(col_idx, ""):
            spotify_col = col_idx
            print(f"✓ Found Spotify column: Col {col_idx} ({headers[col_idx]})")
            break

    if spotify_col:
        # Count non-empty Spotify URLs
        spotify_count = 0
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=spotify_col)
            if cell.value:
                spotify_count += 1
        print(f"  → {spotify_count}/{ws.max_row - 1} rows have Spotify URLs")

    # Check for Era/Period column
    era_col = None
    for col_idx in range(1, ws.max_column + 1):
        header = headers.get(col_idx, "").lower()
        if "תקופה" in headers.get(col_idx, "") or "era" in header or "period" in header:
            era_col = col_idx
            print(f"✓ Found Era column: Col {col_idx} ({headers[col_idx]})")

            # Extract unique values
            eras = set()
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=era_col)
                if cell.value:
                    eras.add(str(cell.value).strip())

            print(f"  → Unique values: {sorted(eras)}")
            break

    # Check for Name column
    name_col = None
    for col_idx in range(1, ws.max_column + 1):
        header = headers.get(col_idx, "").lower()
        if "שם" in headers.get(col_idx, "") or "name" in header:
            name_col = col_idx
            print(f"✓ Found Name column: Col {col_idx} ({headers[col_idx]})")
            break

    print("\n" + "=" * 100)
    print("✅ SUMMARY FOR PYTHON PARSER:")
    print("=" * 100)
    print(f"""
Column Mapping:
  - Name: Column {name_col} ({headers.get(name_col, "?")})
  - Era/Period: Column {era_col} ({headers.get(era_col, "?")})
  - Spotify: Column {spotify_col} ({headers.get(spotify_col, "?")})
  - Max rows: {ws.max_row}
  - Data rows: {ws.max_row - 1}
""")

except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()

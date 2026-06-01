#!/usr/bin/env python3
"""
Full import from Excel to data.json with Spotify queries
"""
import json
import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_FILE = "site-data/חכמי ישראל.xlsx"
OUTPUT_FILE = "data.json"

print("📖 Importing from Excel with Spotify...")

try:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    nodes = []
    skipped = 0

    # Skip header row (row 1)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        try:
            sage_id = str(row[0]) if row[0] else None  # Col 1: ID
            chapter_type = row[1]  # Col 2: סוג פרק (for filtering)
            name = row[2]  # Col 3: שם הדמות
            period_years = row[3]  # Col 4: שנים/תקופה
            location = row[4]  # Col 5: אזור/מרחב
            era = row[5]  # Col 6: תקופה
            field = row[6]  # Col 7: תחום עיקרי
            bio = row[8]  # Col 9: תקציר
            spotify = row[11]  # Col 12: קישור ספוטיפיי

            # Only add if has name and era
            if name and era:
                node = {
                    "id": sage_id or str(len(nodes) + 1),
                    "label": name.strip(),
                    "group": era.lower().replace(' ', '-') if era else 'unknown',
                    "era": era if era else 'Unknown',
                    "period": str(period_years) if period_years else '',
                    "location": str(location).strip() if location else '',
                    "field": str(field).strip() if field else 'Other',
                    "bio": str(bio).strip() if bio else '',
                    "spotifyQuery": str(spotify).strip() if spotify else name
                }
                nodes.append(node)
            else:
                skipped += 1

        except Exception as e:
            print(f"⚠️  Row {row_idx} error: {e}")
            skipped += 1

    print(f"\n✓ Imported {len(nodes)} sages")
    print(f"⚠️  Skipped {skipped} rows (missing name or era)")

    # Save to file
    data = {
        "nodes": nodes,
        "links": []  # Links will be updated manually or via admin panel
    }

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Saved to {OUTPUT_FILE}")
    print(f"\n📊 Sample sages:")
    for i, node in enumerate(nodes[:5]):
        print(f"  {i+1}. {node['label']} ({node['era']})")
        if node['spotifyQuery']:
            print(f"     Spotify: {node['spotifyQuery'][:40]}...")

    print(f"\n🎵 Spotify queries loaded: {sum(1 for n in nodes if n.get('spotifyQuery'))}/{len(nodes)}")

except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()

#!/usr/bin/env python3
"""
CORRECT Excel Parser - Maps exact columns to data.json with proper era mapping
"""
import openpyxl
import json
import sys
import io
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL_FILE = "data/חכמי ישראל.xlsx"
OUTPUT_FILE = "data.json"

# Map Hebrew era names to standardized English + color codes
ERA_MAPPING = {
    # Ancient/Temple periods
    'עת העתיקה': 'Second Temple',
    'ימי בית שני': 'Second Temple',
    'בית שני': 'Second Temple',

    # Tannaim period
    'תנאים': 'Tannaim',
    'תקופת התנאים': 'Tannaim',

    # Amoraim period
    'אמוראים': 'Amoraim',
    'תקופת האמוראים': 'Amoraim',
    'אמוראים ראשונים': 'Amoraim',
    'אמוראים אחרונים': 'Amoraim',

    # Geonim
    'גאונים': 'Geonim',

    # Rishonim period
    'ראשונים': 'Rishonim',
    'תקופת הראשונים': 'Rishonim',
    'ימי הביניים': 'Rishonim',

    # Acharonim period
    'אחרונים': 'Acharonim',
    'תקופת האחרונים': 'Acharonim',

    # Modern
    'עת חדשה': 'Modern',
    'מודרני': 'Modern',
    'עכשוויים': 'Modern',

    # Mixed/transitional
    'תקופת המעבר תנאים–אמוראים': 'Tannaim',
    'חז"ל': 'Amoraim',
    'חז"ל – ימי הביניים': 'Rishonim',
}

print("=" * 100)
print("🔧 PARSING EXCEL WITH CORRECT COLUMNS")
print("=" * 100)

try:
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    nodes = []
    links = []
    skipped = 0
    missing_era = 0
    missing_name = 0
    missing_spotify = 0

    # Parse rows: Col 3 = Name, Col 6 = Era, Col 12 = Spotify
    for row_idx in range(2, ws.max_row + 1):
        try:
            # Read columns (1-indexed in openpyxl)
            sage_id = ws.cell(row=row_idx, column=1).value  # Col 1: ID
            name = ws.cell(row=row_idx, column=3).value  # Col 3: שם
            era_raw = ws.cell(row=row_idx, column=6).value  # Col 6: תקופה
            location = ws.cell(row=row_idx, column=5).value  # Col 5: אזור
            field = ws.cell(row=row_idx, column=7).value  # Col 7: תחום
            bio = ws.cell(row=row_idx, column=9).value  # Col 9: תקציר
            spotify_url = ws.cell(row=row_idx, column=12).value  # Col 12: קישור ספוטיפיי
            related = ws.cell(row=row_idx, column=11).value  # Col 11: דמויות קשורות

            # Skip if no name
            if not name or not str(name).strip():
                missing_name += 1
                continue

            name = str(name).strip()

            # Map era to standardized English name
            era_standardized = None
            if era_raw:
                era_str = str(era_raw).strip()

                # Direct match
                if era_str in ERA_MAPPING:
                    era_standardized = ERA_MAPPING[era_str]
                # Partial match (check if substring)
                else:
                    for key, val in ERA_MAPPING.items():
                        if key in era_str:
                            era_standardized = val
                            break

                # Fallback: try to detect era from year patterns
                if not era_standardized:
                    if 'בית שני' in era_str or 'בנה"ב' in era_str:
                        era_standardized = 'Second Temple'
                    elif 'תנא' in era_str:
                        era_standardized = 'Tannaim'
                    elif 'אמוראים' in era_str:
                        era_standardized = 'Amoraim'
                    elif 'ראשון' in era_str or 'ימי הביניים' in era_str:
                        era_standardized = 'Rishonim'
                    elif 'אחרון' in era_str or 'מודרני' in era_str:
                        era_standardized = 'Acharonim'
                    elif '20' in era_str or '19' in era_str or '21' in era_str:
                        era_standardized = 'Modern'
                    else:
                        era_standardized = 'Unknown'

            if not era_standardized:
                missing_era += 1
                continue

            # Create sage node
            node = {
                "id": str(int(sage_id)) if isinstance(sage_id, (int, float)) else str(len(nodes) + 1),
                "label": name,
                "era": era_standardized,
                "group": era_standardized.lower(),  # lowercase for CSS class
                "period": era_raw if era_raw else '',
                "location": str(location).strip() if location else '',
                "field": str(field).strip() if field else 'Other',
                "bio": str(bio).strip() if bio else f"A sage from {era_standardized}",
                "spotify_url": str(spotify_url).strip() if spotify_url and str(spotify_url).strip() else '',
                "related_sages": str(related).strip() if related else ''
            }

            nodes.append(node)

            if not spotify_url or not str(spotify_url).strip():
                missing_spotify += 1

        except Exception as e:
            print(f"⚠️  Row {row_idx} error: {str(e)[:50]}")
            skipped += 1

    print(f"\n✅ Successfully parsed:")
    print(f"   • Sages: {len(nodes)}")
    print(f"   • With Spotify URLs: {sum(1 for n in nodes if n.get('spotify_url'))}")
    print(f"   • Missing Spotify: {missing_spotify}")
    print(f"   • Missing Era: {missing_era}")
    print(f"   • Missing Name: {missing_name}")

    # Create data structure
    data = {
        "nodes": nodes,
        "links": []
    }

    # Save to file
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Saved to {OUTPUT_FILE}")

    # Print era distribution
    print(f"\n📊 Era Distribution:")
    era_counts = {}
    for node in nodes:
        era = node['era']
        era_counts[era] = era_counts.get(era, 0) + 1

    for era, count in sorted(era_counts.items(), key=lambda x: x[1], reverse=True):
        print(f"   • {era}: {count} sages")

    # Print sample nodes
    print(f"\n📝 Sample nodes (first 3):")
    for i, node in enumerate(nodes[:3]):
        print(f"\n   {i+1}. {node['label']}")
        print(f"      Era: {node['era']}")
        print(f"      Spotify: {node['spotify_url'][:60]}..." if node['spotify_url'] else "      Spotify: [EMPTY]")

except Exception as e:
    print(f"❌ Fatal error: {e}")
    import traceback
    traceback.print_exc()

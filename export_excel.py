#!/usr/bin/env python3
"""
Export 992 sages from Excel to JSON for Supabase import
"""
import openpyxl
import json
import sys
import io

# Fix UTF-8 encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def normalize_period(period_str):
    """Map Excel period names to standard keys"""
    if not period_str:
        return None
    period_str = str(period_str).strip()

    mapping = {
        "עת העתיקה": "second-temple",
        "בית שני": "second-temple",
        "תנאים": "tannaim",
        "ראשונים": "rishonim",
        "אחרונים": "acharonim",
        "עת חדשה": "modern",
    }
    return mapping.get(period_str, period_str)

def parse_related_sages(related_str):
    """Parse comma-separated related sages"""
    if not related_str:
        return []
    return [s.strip() for s in str(related_str).split(',') if s.strip()]

def export_sages():
    wb = openpyxl.load_workbook('site-data/חכמי ישראל.xlsx')
    ws = wb.active

    # Get headers
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            headers[col] = val

    print(f"📊 Found {len(headers)} columns")
    print(f"📈 Processing rows...")

    sages = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col, header in headers.items():
            val = ws.cell(row_idx, col).value
            row_data[header] = val

        # Transform to Supabase schema
        sage = {
            "sage_id": str(row_data.get("מזהה", row_idx)).strip(),
            "name_he": row_data.get("שם הדמות/הנושא", ""),
            "name_en": "",  # Excel doesn't have English names in this sheet
            "chapter_type": row_data.get("סוג פרק", ""),
            "years": row_data.get("שנים/תקופה", ""),
            "area": row_data.get("אזור/מרחב", ""),
            "period_key": normalize_period(row_data.get("תקופה")),
            "main_field": row_data.get("תחום עיקרי", ""),
            "tags": row_data.get("תגיות", ""),
            "summary": row_data.get("תקציר (2–3 שורות)", ""),
            "central_idea": row_data.get("רעיון מרכזי/חידוש", ""),
            "related_sages": ",".join(parse_related_sages(row_data.get("דמויות/השפעות קשורות"))),
            "spotify_link": row_data.get("קישור ספוטיפיי", ""),
            "origin_country": "",
            "migration_path": "",
            "geo_region": "",
            "custom_tradition": "",
        }

        sages.append(sage)

        if row_idx % 100 == 0:
            print(f"  ✓ Processed {row_idx - 1} rows...")

    print(f"✅ Total sages: {len(sages)}")

    # Write to JSON
    with open('sages.json', 'w', encoding='utf-8') as f:
        json.dump(sages, f, ensure_ascii=False, indent=2)

    print(f"💾 Exported to sages.json")
    return len(sages)

if __name__ == "__main__":
    try:
        count = export_sages()
        print(f"\n✨ Done! {count} sages ready for import.")
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)
